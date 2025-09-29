import card_api
import string
from openpyxl import load_workbook, Workbook
from dataclasses import dataclass
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side

@dataclass
class ExcelManager():
    filename: str
    mode: str
        
    def __enter__(self):
        try:
            self.file = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save(filename = self.filename)
            self.file = load_workbook(self.filename)
        return self.file
    
    def __exit__(self, exc_type, exc_value, traceback):
        if ("w" in self.mode or "a" in self.mode):
            try:
                self.file.save(self.filename)
            except PermissionError:
                input("Close Excel and press enter")
                self.__exit__(exc_type, exc_value, traceback)
                return
        
        self.file.close()

def number_to_column(num: int) -> str:
    alphabet = string.ascii_uppercase
    s = ""
    while (num > 0):
        letter_index = ((num - 1) % 26)
        s += alphabet[letter_index]
        num = int((num - 1) / 26)
    return s[::-1]

def column_to_number(col: str) -> int:
    num = 0
    for index,letter in enumerate(col[::-1]):
        num += pow(26, index) * (string.ascii_uppercase.index(letter) + 1)
    return num

def get_first_empty_column(sheet) -> str:
    col_num = 1
    while (sheet[f"{number_to_column(col_num)}1"].value != None):
        col_num += 1
    
    return number_to_column(col_num)

def find_card_in_sheet(card: card_api.Card, sheet: Worksheet) -> int:
    index = 2
    sheet_card = (sheet[f"A{index}"].value, sheet[f"B{index}"].value, sheet[f"C{index}"].value, sheet[f"D{index}"].value)
    card_tuple = (card.name, card.collector_number, card.set, card.foiling)
    
    while True:
        sheet_card = (sheet[f"A{index}"].value, sheet[f"B{index}"].value, sheet[f"C{index}"].value, sheet[f"D{index}"].value)
        if (sheet_card == card_tuple or all(val == None for val in sheet_card)): return index
        index += 1

def compare_tuples(tup1: tuple, tup2: tuple) -> bool:
    if (len(tup1) != len(tup2)): return False
    
    for index in range(len(tup1)):
        if (tup1[index] != tup2[index]): return False
    return True

def set_column_width(sheet: Worksheet, column: str) -> None:
    max_len = 0
    
    for cell in sheet[column]:
        try:
            if (cell.value): max_len = max(max_len, len(str(cell.value)))
        except:
            continue
        
        sheet.column_dimensions[column].width = max_len + 2