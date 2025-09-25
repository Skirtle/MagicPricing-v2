import card_api
import string
from openpyxl import load_workbook, Workbook

def append_to_workbook(wb_filename: str, cards: list[card_api.Card]) -> None:
    try:
        workbook = load_workbook(filename = wb_filename)
    except FileNotFoundError:
        create_workbook(wb_filename)
        append_to_workbook(wb_filename, cards)
        return
    
    print(workbook.sheetnames)
    
def create_workbook(wb_filename: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    
    if (sheet == None): raise ValueError("whar")

    sheet["A1"] = "hello"
    sheet["B1"] = "world!"
    
    workbook.save(filename = wb_filename)
    
def number_to_column(num: int) -> str:
    alphabet = string.ascii_uppercase
    s = ""
    while (num > 0):
        letter_index = ((num - 1) % 26)
        s += alphabet[letter_index]
        num = int((num - 1) / 26)
    return s[::-1]

def column_to_number(col: str) -> int:
    num = 0
    for index,letter in enumerate(col[::-1]):
        num += pow(26, index) * (string.ascii_uppercase.index(letter) + 1)
    return num