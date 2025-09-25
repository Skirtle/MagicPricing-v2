import card_api, logger
import pyodbc, argparse
from os import getcwd
from datetime import datetime
from magic_excel import *
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Border, Side, Font

parser = argparse.ArgumentParser(prog = "Magic Card Pricing", description = "Logs the prices of Magic cards")
parser.add_argument("--dont_read_cache", action = "store_false", default = True, help = "Do not read from prices.cache")
parser.add_argument("--dont_write_cache", action = "store_false", default = True, help = "Do not write to prices.cache")
parser.add_argument("-l", "--log", action = "store_true", default = False, help = "Log events to the log file. NOT IMPLEMENTED YET")
parser.add_argument("-v", "--verbose", action = "store_true", default = False, help = "Print debug information to the screen")
parser.add_argument("-p", "--print_cards", action = "store_true", default = False, help = "Print the cards to screen as the price is found")
parser.add_argument("-V", "--validate", action = "store_true", default = False, help = "Validate the card names. Writes mismatches to validate.txt")
parser.add_argument("-Vo", "--validate_only", action = "store_true", default = False, help = "Same as --validate, but does not continue after validating")
parser.add_argument("--sql", default = "SELECT * FROM Cards", help = "Use a custom SQL query. Default is 'SELECT * FROM Cards'")
parser.add_argument("--log_file", default = "magic.log", help = "The log file's name. NOT YET IMPLEMENTED")
parser.add_argument("--database", default = "Magic.accdb", help = "The database file's read from")
parser.add_argument("--strict_mode", action = "store_true", default = False, help = "When used with --validate, acts as --validate_only")
parser.add_argument("--clear_cache", action = "store_true", default = False, help = "Clear the cache before starting")
parser.add_argument("--excel_filename", default = "magic.xlsx", help = "The filename for the exported Excel spreadsheet. Default = 'magic.xlsx'")
parser.add_argument("-E", "--dont_export", action = "store_true", default = False, help = "Don't export to Excel")

args = parser.parse_args()

def get_cards_from_database(filename: str, sql: str = "", autocall_api: bool = False) -> list[card_api.Card]:
    if (sql == ""): sql = "SELECT * FROM Cards"
    
    # Connect to database
    try:
        encoding = "latin-1"
        directory = getcwd() + "\\"
        driverStr = r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='
        cnxn = pyodbc.connect(driverStr + directory + filename + ";")
        logger.log(f"Connection to {filename} opened", "LOG", args.log_file, args.log, args.verbose)
        cursor = cnxn.cursor()
        cnxn.setdecoding(pyodbc.SQL_CHAR, encoding=encoding)
        cnxn.setdecoding(pyodbc.SQL_WCHAR, encoding=encoding)
        cnxn.setencoding(encoding=encoding)
    except Exception as err:
        print(err)
        return []
    
    # Read the database
    
    try: cursor.execute(sql)
    except Exception as err: cursor.execute("SELECT * FROM Cards")
    
    logger.log(f"Fetching cards using {sql}", "LOG", args.log_file, args.log, args.verbose)
    rows = cursor.fetchall()
    logger.log(f"Found {len(rows)} cards", "LOG", args.log_file, args.log, args.verbose)
    card_list = []
    
    # Create the card objects
    for temp_card in rows:
        card_name = temp_card[0]
        card_cn = temp_card[1]
        card_set = temp_card[2]
        card_foil = temp_card[3].lower()
        card_quantity = int(temp_card[4])
        
        if (card_set == "PLIST"): card_set = "PLST"
        
        card = card_api.Card(card_name, card_cn, card_set, card_foil, quantity = card_quantity, call_api = autocall_api)
        card_list.append(card)
        
    
    # Clean up
    cursor.close()
    cnxn.close()
    
    logger.log(f"Connection to {filename} closed", "LOG", args.log_file, args.log, args.verbose)
    
    return card_list

def read_cards_from_cache() -> dict:
    cache_filename = "prices.cache"
    cache = {}
    today = datetime.today().strftime("%Y%m%d")
    
    with open(cache_filename, "r") as file:
        logger.log(f"Cache file {cache_filename} opened for reading", "LOG", args.log_file, args.log, args.verbose)
        # If date is today, we can continue
        cache_date = file.readline().strip()
        
        if (today == cache_date): # We are good to continue
            prices = [l.strip().split(",") for l in file.readlines()]
            for price in prices: cache[price[0]] = price[1]
            logger.log(f"Found {len(cache)} cached card prices", "LOG", args.log_file, args.log, args.verbose)
        
        else: logger.log(f"Cache is old, ignoring the cache", "WARNING", args.log_file, args.log, args.verbose)
    
    return cache

def write_cards_to_cache(new_cache: dict, old_cache: dict) -> None:
    cache_filename: str = "prices.cache"
    today: str = datetime.today().strftime("%Y%m%d")
    
    cache_date: str = "not today!"
    cards_added: int = 0
    with open(cache_filename, "r") as file: cache_date = file.readline().strip()
    
    if (cache_date != today): 
        # TODO: Rewrite entire cache regardless
        with open(cache_filename, "w") as file:
            logger.log(f"Cache file {cache_filename} opened for writing", "LOG", args.log_file, args.log, args.verbose)
            file.write(f"{today}\n")
            for hash_key in new_cache: 
                cards_added += 1
                file.write(f"{hash_key},{new_cache[hash_key]}\n")
            logger.log(f"Finished writing {cards_added} to the cache", "LOG", args.log_file, args.log, args.verbose)
        
    else:
        # TODO: Check if card is in old_cache. If it is not, append card to file
        with open(cache_filename, "a") as file:
            logger.log(f"Cache file {cache_filename} opened for writing", "LOG", args.log_file, args.log, args.verbose)
            
            for hash_key in new_cache:
                if (hash_key not in old_cache):
                    cards_added += 1
                    file.write(f"{hash_key},{new_cache[hash_key]}\n")
            logger.log(f"Finished writing {cards_added} to the cache", "LOG", args.log_file, args.log, args.verbose)
    
    
    return

def get_card_prices_from_api(cards: list[card_api.Card], check_cache: bool = True, write_to_cache: bool = True) -> bool:
    cache_filename: str = "prices.cache"
    old_cache: dict = {}
    new_cache: dict = {}
    invalid_cards: list[card_api.Card] = []
    today: str = datetime.today().strftime("%Y%m%d")
    
    logger.log(f"Today's date: {today}", "LOG", args.log_file, args.log, args.verbose)
    
    # Make sure the file exists
    file = open(cache_filename, "a")
    file.close()
    
    # Read the cache
    if (check_cache): old_cache = read_cards_from_cache()
    
    # If we are validating, call API for all cards, no matter what
    if (args.validate or args.validate_only): 
        # Call API for all cards, no matter what
        for card in cards:
            card_str = f"{card.name} [{card.set} {card.collector_number} {card.foiling}]"
            print(f"Validation of {logger.Color.BLUE}{card_str}{logger.Color.RESET}", end = "")
            card_valid = validate_card_name(card)
            
            if (not card_valid): 
                # We found an invalid card
                invalid_cards.append(card)
                print(f"{logger.Color.RED} failed {logger.Color.RESET}")
            else:
                print(f"{logger.Color.GREEN} succeeded {logger.Color.RESET}")
        
        
        # Write to file
        with open("validate.txt", "w") as file:
            for card in invalid_cards:
                db_card = f"{card.name} [{card.set} {card.collector_number} {card.foiling}]"
                api_card = f"{card.response_json['name']} [{card.response_json['set'].upper()} {card.response_json['collector_number']}]"
                file.write(f"Got {db_card} but found {api_card}\n")
                
        if ((len(invalid_cards) > 0 and args.strict_mode) or args.validate_only): 
            if (len(invalid_cards) > 0): logger.log("Not all cards succeeded validation, quitting. Check validate.txt", "ERROR", args.log_file, args.log, args.verbose)
            else: logger.log("All cards validated successfully", "LOG", args.log_file, args.log, args.verbose)
            return len(invalid_cards) == 0
        
    # Go through old_cache and get prices
    if (args.print_cards): print("Starting card price fetching")
    for card in cards:
        if (card in invalid_cards): 
            if (args.print_cards): print(f"\tSkipping invalid card {card}")
            continue
        card_hash = card.generate_hash()
        
        if (card_hash in old_cache):
            card.price = float(old_cache[card_hash])
            new_cache[card_hash] = float(old_cache[card_hash])
        
        else:
            # This is where we call the API
            # We will also write the new price to the new_cache, only writing to file if allowed
            card.set_price_from_api()
            new_price = card.price
            new_cache[card_hash] = new_price
        if (args.print_cards): print(f"\tFound {card}")
    
    # Write to cache
    if (write_to_cache): write_cards_to_cache(new_cache, old_cache)
    
    logger.log(f"Finished fetching", "LOG", args.log_file, args.log, args.verbose)
    return len(invalid_cards) == 0

def validate_card_name(card: card_api.Card) -> bool:
    card.response_json = card_api.get_api_response(card) # May as well set the price as well
    return card.name == card.response_json["name"]

def export_excel(filename: str, cards: list[card_api.Card], sheet_name = "Sheet") -> None:
    with ExcelManager(filename, "w") as file:
        sheet = file.active
        center_align = Alignment(horizontal = "center", vertical = "center")
        if (sheet == None): raise ValueError("how")
        sheet.title = sheet_name
        logger.log(f"Opened {filename}, sheet {sheet.title}", "LOG", args.log_file, args.log, args.verbose)
        
        # Initial settings
        sheet["A1"] = "Name"
        sheet["B1"] = "Number"
        sheet["C1"] = "Set"
        sheet["D1"] = "Foiling"
        sheet["E1"] = "Quantity"
        
        # Start writing data in first available space
        # Get the first empty column and set it
        new_column = get_first_empty_column(sheet)
        date_formatted = datetime.now().strftime("%Y-%m-%d")
        sheet[f"{new_column}1"] = date_formatted
        sheet[f"{new_column}1"].number_format = "YYYY-MM-DD"
        
        logger.log(f"Next empty column found: {new_column}, setting date to {date_formatted}", "LOG", args.log_file, args.log, args.verbose)
        
        # Start inputting  card information
        for card in cards:
            card_row = find_card_in_sheet(card, sheet)
            sheet[f"A{card_row}"] = card.name
            sheet[f"B{card_row}"] = card.collector_number
            sheet[f"C{card_row}"] = card.set
            sheet[f"D{card_row}"] = card.foiling
            sheet[f"E{card_row}"] = card.quantity
            sheet[f"{new_column}{card_row}"] = card.price
            sheet[f"{new_column}{card_row}"].number_format = '"$"#,##0.00'
            
        logger.log(f"Adding style and alignment to cells", "LOG", args.log_file, args.log, args.verbose)
        # Center align all of the cells
        center_align = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center_align
                
                
        fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        double_side = Side(border_style="double", color="000000")
        thin_side = Side(border_style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=double_side, bottom=double_side)
        
        # Fix the width of each column and style of each header
        for i in range(1, column_to_number(new_column) + 1):
            set_column_width(sheet, number_to_column(i))
            sheet[f"{number_to_column(i)}1"].fill = fill
            sheet[f"{number_to_column(i)}1"].border = border
            sheet[f"{number_to_column(i)}1"].font = font
            
    logger.log(f"Saved and closed {filename}", "LOG", args.log_file, args.log, args.verbose)

if __name__ == "__main__":
    if (args.clear_cache): 
        file = open("prices.cache", "w")
        file.close()
    
    card_list = get_cards_from_database(args.database)
    cards_valid = get_card_prices_from_api(card_list, args.dont_read_cache, args.dont_write_cache)
    
    if (".xlsx" not in args.excel_filename): excel_filename = args.excel_filename + ".xlsx"
    else: excel_filename = args.excel_filename
    
    if (not args.dont_export): export_excel(excel_filename, card_list, args.database.split(".")[0])
            
    logger.log(f"Done", "LOG", args.log_file, args.log, args.verbose)
            